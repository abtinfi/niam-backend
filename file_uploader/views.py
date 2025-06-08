from rest_framework import generics, permissions, status
from rest_framework.views import APIView
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from django.http import FileResponse
from openpyxl import load_workbook
import matplotlib.pyplot as plt
import os

from .models import UploadedFile, CarbonEntry
from .serializers import UploadedFileSerializer, CarbonEntrySerializer


# -------------------- File Upload --------------------

class FileUploadView(generics.CreateAPIView):
    queryset = UploadedFile.objects.all()
    serializer_class = UploadedFileSerializer
    permission_classes = [permissions.IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


class UserFileListView(generics.ListAPIView):
    serializer_class = UploadedFileSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        return UploadedFile.objects.all() if user.is_staff else UploadedFile.objects.filter(user=user)


# -------------------- Carbon Manual Input --------------------

class CarbonEntryCreateView(generics.CreateAPIView):
    queryset = CarbonEntry.objects.all()
    serializer_class = CarbonEntrySerializer
    permission_classes = [permissions.IsAuthenticated]

    def perform_create(self, serializer):
        serializer.save(user=self.request.user)


# -------------------- Carbon Excel Upload --------------------

class ExcelUploadParseView(APIView):
    parser_classes = [MultiPartParser]
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, *args, **kwargs):
        excel_file = request.FILES.get('file')
        wb = load_workbook(excel_file)
        ws = wb['CarbonData']

        data = {}
        for row in ws.iter_rows(min_row=2, max_row=7):
            key = row[0].value
            value = row[1].value
            data[key] = value

        entry = CarbonEntry.objects.create(
            user=request.user,
            kilometers_per_week=float(data.get('KilometersPerWeek', 0)),
            electricity_per_month=float(data.get('ElectricityPerMonth', 0)),
            short_flights_per_year=int(data.get('ShortFlightsPerYear', 0)),
            long_flights_per_year=int(data.get('LongFlightsPerYear', 0)),
            recycling=str(data.get('Recycling', '')).lower() == 'yes',
            diet_type=str(data.get('DietType', '')).lower()
        )
        return Response({"message": "Data saved", "id": entry.id}, status=status.HTTP_201_CREATED)


# -------------------- Carbon Summary + Tips --------------------

class CarbonSummaryView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        entry = CarbonEntry.objects.get(pk=pk, user=request.user)
        tips = []

        if entry.kilometers_per_week > 100:
            tips.append("Consider reducing car usage or carpooling.")
        if entry.electricity_per_month > 300:
            tips.append("Switch to energy-efficient appliances.")
        if entry.diet_type == "omnivore":
            tips.append("Try reducing red meat consumption.")
        if not entry.recycling:
            tips.append("Start recycling to lower your carbon score.")

        return Response({
            "total_emissions": entry.total_emissions,
            "classification": entry.classification,
            "tips": tips
        })


# -------------------- Carbon Charts --------------------

class CarbonChartView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, pk):
        entry = CarbonEntry.objects.get(pk=pk, user=request.user)

        driving = entry.kilometers_per_week * 52 * 0.21
        electricity = entry.electricity_per_month * 12 * 0.4
        flights = entry.short_flights_per_year * 250 + entry.long_flights_per_year * 1100
        recycling = 0 if entry.recycling else 150
        diet = 1000 if entry.diet_type == 'omnivore' else 600 if entry.diet_type == 'vegetarian' else 300

        labels = ['Driving', 'Electricity', 'Flights', 'Recycling', 'Diet']
        values = [driving, electricity, flights, recycling, diet]

        pie_path = f'/tmp/pie_{pk}.png'
        bar_path = f'/tmp/bar_{pk}.png'

        # Pie Chart
        plt.figure(figsize=(6,6))
        plt.pie(values, labels=labels, autopct='%1.1f%%')
        plt.title('Carbon Footprint Breakdown')
        plt.savefig(pie_path)
        plt.close()

        # Bar Chart
        plt.figure(figsize=(8,5))
        plt.bar(labels, values)
        plt.title('Carbon Emissions by Category (kg/year)')
        plt.ylabel('kg CO₂/year')
        plt.savefig(bar_path)
        plt.close()

        return Response({
            "pie_chart": pie_path,
            "bar_chart": bar_path
        })
